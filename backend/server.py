from fastapi import FastAPI, APIRouter, BackgroundTasks, HTTPException
from fastapi.responses import FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone
import asyncio
import json
import hashlib
import mimetypes
import base64

# Document parsing imports
import PyPDF2
from docx import Document
import openpyxl
import magic
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from PIL import Image
import io

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Global variables for file monitoring
DOCUMENT_ROOT = Path("/var/www/html/MPK/doc/")
file_observer = None
last_scan_time = datetime.now(timezone.utc)

# Document Models
class DocumentFile(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    path: str
    relative_path: str
    file_type: str
    size: int
    content: Optional[str] = None
    thumbnail: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    indexed_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class FileTreeNode(BaseModel):
    name: str
    path: str
    type: str  # 'file' or 'folder'
    children: Optional[List['FileTreeNode']] = []
    file_info: Optional[DocumentFile] = None

class SearchResult(BaseModel):
    document: DocumentFile
    relevance_score: float
    snippet: str

class SearchRequest(BaseModel):
    query: str
    limit: int = 20

# Document content extraction functions
def extract_pdf_content(file_path: Path) -> tuple[str, str]:
    """Extract text content and generate thumbnail from PDF"""
    try:
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
            
            # Generate simple thumbnail (text preview)
            preview_text = text[:200] + "..." if len(text) > 200 else text
            thumbnail = f"data:text/plain;base64,{base64.b64encode(preview_text.encode()).decode()}"
            
            return text.strip(), thumbnail
    except Exception as e:
        logging.error(f"Error extracting PDF content from {file_path}: {e}")
        return "", ""

def extract_docx_content(file_path: Path) -> tuple[str, str]:
    """Extract text content from DOCX"""
    try:
        doc = Document(file_path)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        
        preview_text = text[:200] + "..." if len(text) > 200 else text
        thumbnail = f"data:text/plain;base64,{base64.b64encode(preview_text.encode()).decode()}"
        
        return text.strip(), thumbnail
    except Exception as e:
        logging.error(f"Error extracting DOCX content from {file_path}: {e}")
        return "", ""

def extract_xlsx_content(file_path: Path) -> tuple[str, str]:
    """Extract text content from Excel files"""
    try:
        workbook = openpyxl.load_workbook(file_path)
        text = ""
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"Sheet: {sheet_name}\n"
            
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
        
        preview_text = text[:200] + "..." if len(text) > 200 else text
        thumbnail = f"data:text/plain;base64,{base64.b64encode(preview_text.encode()).decode()}"
        
        return text.strip(), thumbnail
    except Exception as e:
        logging.error(f"Error extracting Excel content from {file_path}: {e}")
        return "", ""

def extract_text_content(file_path: Path) -> tuple[str, str]:
    """Extract content from plain text files"""
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            text = file.read()
        
        preview_text = text[:200] + "..." if len(text) > 200 else text
        thumbnail = f"data:text/plain;base64,{base64.b64encode(preview_text.encode()).decode()}"
        
        return text, thumbnail
    except Exception as e:
        logging.error(f"Error extracting text content from {file_path}: {e}")
        return "", ""

def extract_document_content(file_path: Path) -> tuple[str, str]:
    """Extract content based on file type"""
    try:
        mime_type, _ = mimetypes.guess_type(str(file_path))
        file_extension = file_path.suffix.lower()
        
        if file_extension == '.pdf' or mime_type == 'application/pdf':
            return extract_pdf_content(file_path)
        elif file_extension in ['.docx', '.doc'] or mime_type in ['application/vnd.openxmlformats-officedocument.wordprocessingml.document']:
            return extract_docx_content(file_path)
        elif file_extension in ['.xlsx', '.xls'] or mime_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
            return extract_xlsx_content(file_path)
        elif file_extension in ['.txt', '.rtf'] or mime_type == 'text/plain':
            return extract_text_content(file_path)
        else:
            return "", ""
    except Exception as e:
        logging.error(f"Error determining file type for {file_path}: {e}")
        return "", ""

async def index_file(file_path: Path) -> Optional[DocumentFile]:
    """Index a single file"""
    try:
        if not file_path.exists() or file_path.is_dir():
            return None
        
        # Extract content
        content, thumbnail = extract_document_content(file_path)
        
        if not content:  # Skip files we can't parse
            return None
        
        relative_path = str(file_path.relative_to(DOCUMENT_ROOT))
        file_stats = file_path.stat()
        
        doc_file = DocumentFile(
            name=file_path.name,
            path=str(file_path),
            relative_path=relative_path,
            file_type=file_path.suffix.lower(),
            size=file_stats.st_size,
            content=content,
            thumbnail=thumbnail
        )
        
        # Store in database
        doc_dict = doc_file.dict()
        doc_dict['created_at'] = doc_dict['created_at'].isoformat()
        doc_dict['updated_at'] = doc_dict['updated_at'].isoformat()
        doc_dict['indexed_at'] = doc_dict['indexed_at'].isoformat()
        
        await db.documents.replace_one(
            {"path": str(file_path)}, 
            doc_dict, 
            upsert=True
        )
        
        logging.info(f"Indexed file: {file_path}")
        return doc_file
        
    except Exception as e:
        logging.error(f"Error indexing file {file_path}: {e}")
        return None

async def scan_directory():
    """Scan the document directory and index all files"""
    logging.info("Starting directory scan...")
    
    if not DOCUMENT_ROOT.exists():
        logging.warning(f"Document root {DOCUMENT_ROOT} does not exist")
        return
    
    supported_extensions = {'.pdf', '.docx', '.doc', '.xlsx', '.xls', '.txt', '.rtf'}
    
    for file_path in DOCUMENT_ROOT.rglob("*"):
        if file_path.is_file() and file_path.suffix.lower() in supported_extensions:
            await index_file(file_path)
    
    global last_scan_time
    last_scan_time = datetime.now(timezone.utc)
    logging.info("Directory scan completed")

def build_file_tree(path: Path) -> FileTreeNode:
    """Build file tree structure"""
    node = FileTreeNode(
        name=path.name,
        path=str(path),
        type='folder' if path.is_dir() else 'file',
        children=[]
    )
    
    if path.is_dir():
        try:
            for child in sorted(path.iterdir()):
                if child.is_dir() or child.suffix.lower() in {'.pdf', '.docx', '.doc', '.xlsx', '.xls', '.txt', '.rtf'}:
                    node.children.append(build_file_tree(child))
        except PermissionError:
            pass
    
    return node

# Background task for periodic scanning
async def periodic_scan():
    """Periodic scanning task"""
    while True:
        try:
            await asyncio.sleep(300)  # 5 minutes
            await scan_directory()
        except Exception as e:
            logging.error(f"Error in periodic scan: {e}")

# API Routes
@api_router.get("/")
async def root():
    return {"message": "Document Search API", "status": "running"}

@api_router.post("/scan")
async def manual_scan(background_tasks: BackgroundTasks):
    """Manually trigger directory scan"""
    background_tasks.add_task(scan_directory)
    return {"message": "Scan initiated"}

@api_router.get("/file-tree")
async def get_file_tree():
    """Get file tree structure"""
    if not DOCUMENT_ROOT.exists():
        return {"error": "Document root not found"}
    
    tree = build_file_tree(DOCUMENT_ROOT)
    return tree

@api_router.post("/search")
async def search_documents(request: SearchRequest):
    """Search documents"""
    try:
        if not request.query.strip():
            return {"results": [], "total": 0}
        
        # MongoDB text search
        pipeline = [
            {
                "$match": {
                    "$text": {"$search": request.query}
                }
            },
            {
                "$addFields": {
                    "relevance_score": {"$meta": "textScore"}
                }
            },
            {
                "$sort": {"relevance_score": {"$meta": "textScore"}}
            },
            {"$limit": request.limit}
        ]
        
        # Create text index if it doesn't exist
        try:
            await db.documents.create_index([("content", "text"), ("name", "text")])
        except Exception as e:
            logging.info(f"Text index already exists or error creating: {e}")
        
        results = []
        async for doc in db.documents.aggregate(pipeline):
            # Generate snippet
            content = doc.get('content', '')
            query_lower = request.query.lower()
            
            # Find best snippet
            snippet_start = content.lower().find(query_lower)
            if snippet_start != -1:
                start = max(0, snippet_start - 50)
                end = min(len(content), snippet_start + len(request.query) + 100)
                snippet = "..." + content[start:end] + "..."
            else:
                snippet = content[:150] + "..." if len(content) > 150 else content
            
            # Parse datetime strings back to datetime objects
            if isinstance(doc.get('created_at'), str):
                doc['created_at'] = datetime.fromisoformat(doc['created_at'])
            if isinstance(doc.get('updated_at'), str):
                doc['updated_at'] = datetime.fromisoformat(doc['updated_at'])
            if isinstance(doc.get('indexed_at'), str):
                doc['indexed_at'] = datetime.fromisoformat(doc['indexed_at'])
            
            search_result = SearchResult(
                document=DocumentFile(**doc),
                relevance_score=doc.get('relevance_score', 0),
                snippet=snippet
            )
            results.append(search_result)
        
        return {"results": results, "total": len(results), "query": request.query}
        
    except Exception as e:
        logging.error(f"Search error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/documents/{doc_id}")
async def get_document(doc_id: str):
    """Get a specific document"""
    doc = await db.documents.find_one({"id": doc_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    # Parse datetime strings
    if isinstance(doc.get('created_at'), str):
        doc['created_at'] = datetime.fromisoformat(doc['created_at'])
    if isinstance(doc.get('updated_at'), str):
        doc['updated_at'] = datetime.fromisoformat(doc['updated_at'])
    if isinstance(doc.get('indexed_at'), str):
        doc['indexed_at'] = datetime.fromisoformat(doc['indexed_at'])
    
    return DocumentFile(**doc)

@api_router.get("/documents/{doc_id}/download")
async def download_document(doc_id: str):
    """Download a specific document"""
    doc = await db.documents.find_one({"id": doc_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    file_path = Path(doc['path'])
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found on disk")
    
    return FileResponse(
        path=str(file_path),
        filename=doc['name'],
        media_type='application/octet-stream'
    )

@api_router.get("/documents/path/{path:path}")
async def get_document_by_path(path: str):
    """Get a document by its relative path"""
    doc = await db.documents.find_one({"relative_path": path})
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    # Parse datetime strings
    if isinstance(doc.get('created_at'), str):
        doc['created_at'] = datetime.fromisoformat(doc['created_at'])
    if isinstance(doc.get('updated_at'), str):
        doc['updated_at'] = datetime.fromisoformat(doc['updated_at'])
    if isinstance(doc.get('indexed_at'), str):
        doc['indexed_at'] = datetime.fromisoformat(doc['indexed_at'])
    
    return DocumentFile(**doc)

@api_router.get("/stats")
async def get_stats():
    """Get indexing statistics"""
    total_docs = await db.documents.count_documents({})
    return {
        "total_documents": total_docs,
        "last_scan": last_scan_time.isoformat(),
        "document_root": str(DOCUMENT_ROOT)
    }

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup_event():
    """Initialize the application"""
    # Create initial index
    asyncio.create_task(scan_directory())
    # Start periodic scanning
    asyncio.create_task(periodic_scan())
    logging.info("Document indexer started")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()